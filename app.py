from flask import Flask, jsonify, request
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
from sqlalchemy.exc import SQLAlchemyError

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///employees.db'
db = SQLAlchemy(app)

#Database models
class Company(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    employees = db.relationship('Employee', backref='company')

class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    phone_number = db.Column(db.String(20), unique=True, nullable=False)
    salary = db.Column(db.Float, nullable=False)
    manager_id = db.Column(db.Integer, db.ForeignKey('employee.id'))
    department_id = db.Column(db.Integer)
    company_id = db.Column(db.Integer, db.ForeignKey('company.id'))
    
with app.app_context():
    db.create_all()

#Route to load data from excel file
@app.route('/load_data', methods=['POST'])
def load_data():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file part'}), 400

        excel_file = request.files['file']
        
        if excel_file.filename == '':
            return jsonify({'error': 'No selected file'}), 400
        
        workbook = load_workbook(excel_file)
        worksheet = workbook.active

        companies = {}
        employees = []

        #Bulk upload to avoid recurring queries of same data
        for row in worksheet.iter_rows(values_only=True, min_row=2):
            employee_id, first_name, last_name, phone_number, company_name, salary, manager_id, department_id = row

            if company_name not in companies:
                try:
                    company = Company(name=company_name)
                    db.session.add(company)
                    db.session.commit()
                    companies[company_name] = company.id
                except SQLAlchemyError as e:
                    db.session.rollback()
                    return jsonify({'error': f'Error adding company: {str(e)}'}), 500

            employee = Employee(
                id=employee_id,
                first_name=first_name,
                last_name=last_name,
                phone_number=phone_number,
                salary=salary,
                manager_id=manager_id,
                department_id=department_id,
                company_id=companies[company_name]
            )
            employees.append(employee)

        try:
            db.session.add_all(employees)
            db.session.commit()
        except SQLAlchemyError as e:
            db.session.rollback()
            return jsonify({'error': f'Error adding employees: {str(e)}'}), 500

        return jsonify({'message': 'Data loaded successfully'})

    except Exception as e:
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True)
